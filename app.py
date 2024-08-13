import os
import re
import time
import requests
from urllib import parse
from io import BytesIO

import cv2
import numpy as np
import PIL
from PIL import Image
import openpyxl
import xlrd
from xpinyin import Pinyin
from fake_useragent import UserAgent
from tqdm import tqdm


def create_dir_or_file(path):
    if not os.path.exists(path):
        file_name, file_extension = os.path.splitext(path)
        if file_extension == "":
            os.makedirs(path, exist_ok=True)
        else:
            if not os.path.exists(path):
                with open(path, 'w') as file:
                    pass

def read(pic_path):
    img = cv2.imread(pic_path)
    if img is None:
        return 0
    W, H = img.shape[1], img.shape[0]
    return img, W, H

def res_detect(W, H):
    if W < resolution or H < resolution:
        print("0 res_detect | W, H: ", W, H)
        return False
    else:
        print("1 res_detect | W, H: ", W, H)
        return True

def face_area_ok(face_area, img_area):
    face_rate = face_area / img_area
    if face_rate < face_img_rate:
        print(f"0 face_area_ok | {face_area} / {img_area} = {face_rate} < {face_img_rate}")
        return False
    else:
        print(f"1 face_area_ok | {face_area} / {img_area} = {face_rate} >= {face_img_rate}")
        return True

def crop(img, H, W, center):
    max_size = min(H, W)
    
    l_x, l_y = 0, 0

    if W > H:
        l_y = 0
        if center[0] < max_size // 2:
            l_x = 0
        elif center[0] > W - max_size // 2:
            l_x = W - max_size
        else:
            l_x = center[0] - max_size // 2
    elif H > W:
        l_x = 0
        if center[1] < max_size // 2:
            l_y = 0
        elif center[1] > H - max_size // 2:
            l_y = H - max_size
        else:
            l_y = center[1] - max_size // 2

    img = img[l_y:l_y + max_size, l_x:l_x + max_size]
    return img

def resize(img, img_size):
    img = cv2.resize(img, (int(img_size), int(img_size)))
    return img

def read_excel(excel_path):
    book = xlrd.open_workbook(excel_path)
    sheet = book.sheet_by_name(sheet_name)
    nrows, ncols = sheet.nrows, sheet.ncols

    name_list = []
    for i in range(nrows):
        row = sheet.row_values(i)
        name_list.append(row)

    return name_list

def number_to_letter(num):
    letter = chr(64 + num)
    return letter

def cn2en_write(excel_path):
    book = xlrd.open_workbook(excel_path)
    sheet = book.sheet_by_name(sheet_name)
    cn_names = sheet.col_values(0)
    cn_names = list(set([cn_name for cn_name in cn_names if cn_name != ""]))
    print("cn_names", cn_names, len(cn_names))

    p = Pinyin()
    cn_en_name_list = []
    for i in range(len(cn_names)):
        result1 = p.get_pinyin(cn_names[i])
        s = result1.split('-')
        result3 = s[0].capitalize() + ''.join(s[1:]).capitalize()
        cn_en_name_list.append([cn_names[i], result3, 0])
    cn_en_name_list = sorted(cn_en_name_list, key=lambda x: x[1])
    print('cn_en_name_list', cn_en_name_list)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    # Set to blank firstly
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    
    for r in range(len(cn_en_name_list)):
        for c in range(len(cn_en_name_list[0])):
            ws[f"{number_to_letter(c + 1)}" + str(r + 1)] = cn_en_name_list[r][c]
    wb.save(excel_path)

    return cn_en_name_list

class Spider:
    def __init__(self, I, cn_name, en_name, en_name_dir):
        self.I = I
        self.cn_name = cn_name
        self.en_name = en_name
        self.en_name_dir = en_name_dir
        self.name = parse.quote(self.cn_name + limit)
        self.times = str(int(time.time() * 1000))

        self.url = None
        if size_type == "all": # 全部尺寸
            self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=8032920601831512061&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=&height=&face=&istype=&qc=&nc=1&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        elif size_type == "extra large": # 特大尺寸
            self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=5314417940526052016&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=9&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=0&height=0&face=&istype=&qc=&nc=&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        else: # 未指定尺寸，则默认用全部尺寸
            self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=8032920601831512061&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=&height=&face=&istype=&qc=&nc=1&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        
        self.headers = {'User-Agent': UserAgent().random}

    # 请求30张图片的链接
    def get_one_html(self, url, pn):
        while True:
            try:
                response = requests.get(url=url.format(self.name, self.name, pn, self.times), headers=self.headers).content.decode('utf-8')
                break
            except Exception as e:
                print(f"error {e} occurred. Retrying...")
                time.sleep(1)
        return response

    # 请求单张图片内容
    def get_two_html(self, url):
        while True:
            try:
                response = requests.get(url=url, headers=self.headers).content
                break
            except Exception as e:
                print(f"error {e} occurred. Retrying...")
                time.sleep(1)
        return response

    # 解析含30张图片的html的内容
    def parse_html(self, regex, html):
        content = regex.findall(html)
        return content

    def well_detection(self, img, pic_path):
        # 读取图片
        W, H = img.shape[1], img.shape[0]
        print('W,H读取没问题')

        # 判断分辨率, 如果低于resolution，返回0
        _res_detect = res_detect(W, H)
        if not _res_detect:
            return 0

        # 检测人脸
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        if len(faces) >= 1:
            print("1 faces | len(faces): ", len(faces))
            for (x, y, w, h) in faces:
                print('Dealing with circular face: ', pic_path)
                # 不需要画上框格
                # img=cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)

                # Define the region of interest in the image  
                # img = img[y:y+h, x:x+w]

                # l_x,l_y=x-int((w//2)*(b-1)),y-int((h//2)*(b-1))
                # r_x,r_y=x+w+int((w//2)*(b-1)),y+h+int((h//2)*(b-1))
                # if l_x<0:l_x=0
                # if l_y<0:l_y=0
                # if r_x>H:r_x=H
                # if r_y>W:r_y>W
                
                # 判断人脸大小是否满足比例
                face_area, img_area = w*h, H*W
                _face_area_ok = face_area_ok(face_area, img_area)
                if not _face_area_ok:
                    return 0
                else:
                    # 开始裁剪
                    center = [(2 * x + w) // 2, (2 * y + h) // 2]
                    img = crop(img, H, W, center)

                    # 调整大小
                    img = resize(img, img_size)

                    # 重新保存图片
                    cv2.imwrite(pic_path, img)
                    
                    return 1
                    
        else:
            print("0 faces | len(faces): ", len(faces))
            return 0

    def run(self):
        # #判断该目录下是否存在与输入名称一样的文件夹 如果没有则创建 有就不执行if下的创建
        # if not os.path.exists('./{}/'.format(self.en_name_dir)):
        #     os.mkdir('./{}'.format(self.en_name_dir))
        print('self.url: ', self.url)
        response = self.get_one_html(self.url, 0)
        regex1 = re.compile('"displayNum":(.*?),')
        num = self.parse_html(regex1, response)[0] #获取总的照片数量
        print('{}{}一共有{}张照片'.format(self.cn_name, self.en_name,num)) #打印总的照片数量

        #判断总数能不能整除30
        if int(num) % 30 == 0:
            pn = int(num) / 30
        else:
            # 总数量除30是因为每一个链接有30张照片 +2是因为要想range最多取到该数就需要+1
            # 另外的+1是因为该总数除30可能有余数，有余数就需要一个链接 所以要+1
            pn = int(num) // 30 + 2

        num_pic = 0
        for i in range(int(pn)): #遍历每一个含30张图片的链接
            resp = self.get_one_html(self.url, i * 30)
            urls = re.findall('"ObjURL":"(.*?)",', resp, re.S)
            print("urls1",urls)
            for i in range(len(urls)):
                urls[i] = ''.join(urls[i].split('\\'))
            print("urls2", urls)

            # resp = self.get_one_html(self.url, i * 30)
            # regex2 = re.compile('"middleURL":"(.*?)"')
            # urls = self.parse_html(regex2,resp) #得到30张图片的链接（30个）

            for j, u in enumerate(urls):  # 遍历每张图片的链接
                print('u:', u)
                if u.split(':')[0] == 'https': # and (u[-3:]=='jpg' or u[-4:]=='jpeg'):
                    try:
                        response = requests.get(url=u, headers=self.headers, timeout=(20.00, 10.00))
                        content = response.content
                        # content = self.get_two_html(u) #请求每张图片的内容
                    except requests.exceptions.ConnectionError as e:
                        print('ConnectionError:', e)
                        continue
                    except requests.exceptions.ReadTimeout as e:
                        print('ReadTimeout:', e)
                        continue
                    except requests.exceptions.ChunkedEncodingError as e:
                        print('ChunkedEncodingError:', e)
                        continue
                    # content = self.get_two_html(u) #请求每张图片的内容

                    print('下面要打开了')
                    try:
                        img = Image.open(BytesIO(content)).convert("RGB")
                    except PIL.UnidentifiedImageError as e:
                        print('PIL.UnidentifiedImageError:', e)
                        continue
                    except OSError as e:
                        print('OSError:', e)
                        continue
                    img = cv2.cvtColor(np.asarray(img), cv2.COLOR_RGB2BGR)

                    pic_path = os.path.join(root_dir, sheet_name, self.en_name_dir, star_concept_name, f'{num_pic}.jpg')
                    # pic_path='{}/{}/{}/my_concept/{}.jpg'.format(root_dir,sheet_name,self.en_name_dir,num_pic)
                    # with open(pic_path,'wb') as f:
                    #     f.write(content)
                    #     print('爬保:',pic_path)
                    
                    # 看是否满足条件
                    res = self.well_detection(img, pic_path)

                    if res == 1:
                        num_pic += 1
                        print('saved 第{}人{}{}, 第{}/{}张照片, pic_path:{}\n'.format(self.I, self.cn_name, self.en_name, num_pic, max_pic_num, pic_path))
                    else:
                        print('不满足条件\n')
                        # os.remove(pic_path)
                        # print('删除了: 第{}张照片：{}\n'.format(num_pic,pic_path))

                if num_pic == max_pic_num:
                    break
            if num_pic == max_pic_num:
                break



# 基本参数设置
root_dir = 'output'
# is_man = True
# 每个人需要下载多少张图片
max_pic_num = 10
# 从第几个人开始 第一行就是序号为0
start = 0
# 指定下载图片尺寸类型
size_type = "extra large"
# 放大倍数
# b = 1.5
# 裁剪后可接受的长宽比
# rate = 1.3
#' 正脸 高清 大尺寸'
limit = ''
# 分别对应 全部尺寸0 特大尺寸9 大尺寸3 中尺寸2 小尺寸1 暂时没用到该参数 还是调上面self.url就行了
size_level = 9
# 分辨率分界线，低于这个的pass掉
resolution = 768
# 脸图比例，小于这个不考虑
face_img_rate = 0.03
# resize后的图像大小
img_size = 768

# 根据基本参数变化的间接参数
sheet_name = 'Sheet_man' # if is_man else 'woman'
excel_path = 'person_names.xlsx'
# sheet_num = 1 # if is_man else 2
star_concept_name = 'concept'
# star_output_name = 'generated'

def main():
    create_dir_or_file(root_dir)
    # 中文名转为英文名，写入文件
    cn2en_write(excel_path)

    # 获取中英名list:[[cn,en],[cn,en],...]
    name_list = read_excel(excel_path)
    print('name_list: ', name_list)

    for I in tqdm(range(start, len(name_list))):
        cn_name, en_name = name_list[I][0], name_list[I][1]
        star_concept_dir = os.path.join(root_dir, sheet_name, en_name, star_concept_name)
        # star_output_dir = os.path.join(root_dir, sheet_name, en_name, star_output_name)
        if os.path.exists(star_concept_dir):
            if os.listdir(star_concept_dir) == max_pic_num:
                continue
        else:
            create_dir_or_file(star_concept_dir)
            # create_dir_or_file(star_output_dir)

            print('Downloading images: {}/{}: {} {}'.format(I, len(name_list), cn_name, en_name))
            spider = Spider(I, cn_name, en_name, en_name)
            spider.run()

if __name__ == '__main__':
    main()
