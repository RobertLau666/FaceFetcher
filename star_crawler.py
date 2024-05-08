# -!- coding: utf-8 -!-
import re
import requests
import time
from urllib import parse
import os
import cv2
import PIL
from PIL import Image
from io import BytesIO
import numpy as np

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import ExcelWriter
import xlrd
from xpinyin import Pinyin

from fake_useragent import UserAgent
from tqdm.auto import tqdm

# envpath = '/data/storage1/public/chenyu.liu/anaconda3/envs/py37/lib/python3.7/site-packages/cv2/qt/plugins/platforms'
# os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = envpath

def create_dir_or_file(path):
    if not os.path.exists(path):
        file_name, file_extension = os.path.splitext(path)
        if file_extension == "":
            os.makedirs(path, exist_ok=True)
        else:
            if not os.path.exists(path):
                with open(path, 'w') as file:
                    pass

def read(pic_path):
    img=cv2.imread(pic_path)
    if img is None:
        return 0
    W,H = img.shape[1],img.shape[0]
    return img,W,H

def res_detect(W,H):
    if W<resolution or H<resolution:
        print("0 res_detect | W,H:",W,H)
        return False
    else:
        print("1 res_detect | W,H:",W,H)
        return True

def face_area_ok(face_area,img_area):
    face_rate=face_area/img_area
    if face_rate<face_img_rate:
        print(f"0 face_area_ok | {face_area}/{img_area}={face_rate}<{face_img_rate}")
        return False
    else:
        print(f"1 face_area_ok | {face_area}/{img_area}={face_rate}>={face_img_rate}")
        return True

def crop(img,H,W,center):
    max_size=min(H,W)
    
    l_x,l_y=0,0

    if W>H:
        l_y=0
        if center[0]<max_size//2:
            l_x=0
        elif center[0]>W-max_size//2:
            l_x=W-max_size
        else:
            l_x=center[0]-max_size//2
    elif H>W:
        l_x=0
        if center[1]<max_size//2:
            l_y=0
        elif center[1]>H-max_size//2:
            l_y=H-max_size
        else:
            l_y=center[1]-max_size//2

    img = img[l_y:l_y+max_size, l_x:l_x+max_size]

    print("croped",img.shape)
    return img

def resize(img,img_size):
    img = cv2.resize(img, (int(img_size), int(img_size)))

    print("resized",img.shape)
    return img

class Picture:
    def __init__(self,I,cn_name,en_name,en_name_dir):
        self.I=I
        self.cn_name = cn_name
        self.en_name = en_name
        self.en_name_dir=en_name_dir
        self.name = parse.quote(self.cn_name+limit) #周杰伦 --> 编码
        self.times = str(int(time.time()*1000)) #时间戳-->补全url
        # self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=8032920601831512061&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=&height=&face=&istype=&qc=&nc=1&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        # 全部尺寸
        # self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=8032920601831512061&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=&height=&face=&istype=&qc=&nc=1&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        # 特大尺寸
        # self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=5314417940526052016&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=9&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=0&height=0&face=&istype=&qc=&nc=&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        
        self.url = None
        if size_type == "all": # 全部尺寸
            self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=8032920601831512061&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=&height=&face=&istype=&qc=&nc=1&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        elif size_type == "extra large": # 特大尺寸
            self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=5314417940526052016&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=9&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=0&height=0&face=&istype=&qc=&nc=&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        else: # 未指定尺寸
            self.url = 'https://image.baidu.com/search/acjson?tn=resultjson_com&logid=8032920601831512061&ipn=rj&ct=201326592&is=&fp=result&fr=&word={}&cg=star&queryWord={}&cl=2&lm=-1&ie=utf-8&oe=utf-8&adpicid=&st=&z=&ic=&hd=&latest=&copyright=&s=&se=&tab=&width=&height=&face=&istype=&qc=&nc=1&expermode=&nojc=&isAsync=&pn={}&rn=30&gsm=1e&{}='
        
        self.headers = {'User-Agent':UserAgent().random}

    #请求30张图片的链接
    def get_one_html(self,url,pn):
        response = requests.get(url=url.format(self.name,self.name, pn, self.times), headers=self.headers).content.decode('utf-8')
        return response

    #请求单张图片内容
    def get_two_html(self,url):
        response = requests.get(url=url, headers=self.headers).content
        return response

    #解析含30张图片的html的内容
    def parse_html(self,regex,html):
        content = regex.findall(html)
        return content

    def well_detection(self,img,pic_path):
        # 检测人脸数量
        # face_num=get_face_num(content)
        # if face_num==1:
        #     # 裁剪、调整大小
        #     content=crop_resize(content)
        #     return content
        # else:
        #     return False

        # 读取图片
        # img,W,H=read(pic_path)
        W,H = img.shape[1],img.shape[0]
        print('W,H读取没问题')

        # 判断分辨率, 如果低于resolution，返回0
        _res_detect=res_detect(W,H)
        if not _res_detect:
            return 0

        # 检测人脸
        face_cascade=cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        faces=face_cascade.detectMultiScale(gray,1.3,5)

        if len(faces)>=1:
            print("1 faces | len(faces):",len(faces))
            for (x,y,w,h) in faces:
                print('正在处理循环脸：',pic_path)
                # 不需要画上框格
                # img=cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)

                # Define the region of interest in the image  
                # img = img[y:y+h, x:x+w]

                # l_x,l_y=x-int((w//2)*(b-1)),y-int((h//2)*(b-1))
                # r_x,r_y=x+w+int((w//2)*(b-1)),y+h+int((h//2)*(b-1))
                # if l_x<0:l_x=0
                # if l_y<0:l_y=0
                # if r_x>H:r_x=H
                # if r_y>W:r_y>W
                
                # 判断人脸大小是否满足比例
                face_area,img_area=w*h,H*W
                _face_area_ok=face_area_ok(face_area,img_area)
                if not _face_area_ok:
                    return 0
                else:
                    # 开始裁剪
                    center=[(2*x+w)//2,(2*y+h)//2]
                    img=crop(img,H,W,center)

                    # 调整大小
                    img=resize(img,img_size)

                    # 重新保存图片
                    cv2.imwrite(pic_path,img)
                    
                    return 1
                    
        else:
            print("0 faces | len(faces):",len(faces))
            return 0


    #主函数
    def run(self):
        # #判断该目录下是否存在与输入名称一样的文件夹 如果没有则创建 有就不执行if下的创建
        # if not os.path.exists('./{}/'.format(self.en_name_dir)):
        #     os.mkdir('./{}'.format(self.en_name_dir))
        print('self.url:',self.url)
        response = self.get_one_html(self.url,0)
        regex1 = re.compile('"displayNum":(.*?),')
        num = self.parse_html(regex1,response)[0] #获取总的照片数量
        print('{}{}一共有{}张照片'.format(self.cn_name,self.en_name,num)) #打印总的照片数量

        #判断总数能不能整除30
        if int(num)%30 == 0:
            pn = int(num)/30
        else:
            # 总数量除30是因为每一个链接有30张照片 +2是因为要想range最多取到该数就需要+1
            # 另外的+1是因为该总数除30可能有余数，有余数就需要一个链接 所以要+1
            pn = int(num)//30 + 2

        num_pic=0
        for i in range(int(pn)): #遍历每一个含30张图片的链接
            resp = self.get_one_html(self.url, i * 30)
            urls = re.findall('"ObjURL":"(.*?)",', resp, re.S)
            print("urls1",urls)
            for i in range(len(urls)):
                urls[i]=''.join(urls[i].split('\\'))
            print("urls2",urls)

            # resp = self.get_one_html(self.url, i * 30)
            # regex2 = re.compile('"middleURL":"(.*?)"')
            # urls = self.parse_html(regex2,resp) #得到30张图片的链接（30个）

            for j,u in enumerate(urls):  #遍历每张图片的链接
                print('u:',u)
                if u.split(':')[0]=='https':# and (u[-3:]=='jpg' or u[-4:]=='jpeg'):
                    try:
                        response = requests.get(url=u, headers=self.headers,timeout=(20.00, 10.00))
                        content = response.content
                        # content = self.get_two_html(u) #请求每张图片的内容
                    except requests.exceptions.ConnectionError as e:
                        print('抛出异常 ConnectionError:', e)
                        continue
                    except requests.exceptions.ReadTimeout as e:
                        print('抛出异常 ReadTimeout:', e)
                        continue
                    except requests.exceptions.ChunkedEncodingError as e:
                        print('抛出异常 ChunkedEncodingError:', e)
                        continue
                    # content = self.get_two_html(u) #请求每张图片的内容

                    print('下面要打开了')
                    try:
                        img = Image.open(BytesIO(content)).convert("RGB")
                    except PIL.UnidentifiedImageError as e:
                        print('抛出异常 PIL.UnidentifiedImageError:', e)
                        continue
                    except OSError as e:
                        print('抛出异常 OSError:', e)
                        continue
                    img = cv2.cvtColor(np.asarray(img), cv2.COLOR_RGB2BGR)

                    pic_path=os.path.join(root_dir,gender,self.en_name_dir,star_concept_name,f'{num_pic}.jpg')
                    # pic_path='{}/{}/{}/my_concept/{}.jpg'.format(root_dir,gender,self.en_name_dir,num_pic)
                    # with open(pic_path,'wb') as f:
                    #     f.write(content)
                    #     print('爬保:',pic_path)
                    
                    # 看是否满足条件
                    res=self.well_detection(img,pic_path)

                    if res==1:
                        num_pic+=1
                        print('saved 第{}人{}{},第{}/{}张照片,pic_path:{}\n'.format(self.I,self.cn_name,self.en_name,num_pic,max_pic_num,pic_path)) #下载完一张图片后打印
                    else:
                        print('不满足条件\n')
                        # os.remove(pic_path)
                        # print('删除了: 第{}张照片：{}\n'.format(num_pic,pic_path))

                if num_pic==max_pic_num:
                    break
            if num_pic==max_pic_num:
                break

def read_excel(excel_path):
    book = xlrd.open_workbook(excel_path)

    sheet = book.sheet_by_name(u'Sheet{}'.format(sheet_num))  # 通过名称获取 u表示后面字符串以 Unicode 格式 进行编码，一般用在中文字符串前面，以防乱码

    nrows = sheet.nrows
    ncols = sheet.ncols

    name_list =[]
    for i in range(nrows):
        row = sheet.row_values(i)
        name_list.append(row)

    return name_list


def cn_en_write(excel_path):
    book = xlrd.open_workbook(excel_path)
    sheet = book.sheet_by_name(u'Sheet{}'.format(sheet_num))  # 通过名称获取 u表示后面字符串以 Unicode 格式 进行编码，一般用在中文字符串前面，以防乱码
    cn_name_list = sheet.col_values(0)

    p = Pinyin()
    cn_en_name_list = []
    for i in range(len(cn_name_list)):
        result1 = p.get_pinyin(cn_name_list[i])
        s = result1.split('-')
        result3 = s[0].capitalize() + ''.join(s[1:]).capitalize()
        # result3=''.join(result3.split())
        cn_en_name_list.append([cn_name_list[i], result3])
    cn_en_name_list = sorted(cn_en_name_list, key=lambda x: x[1])
    print('cn_en_name_list', cn_en_name_list)
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sheet{}".format(sheet_num)]
    for i in range(len(cn_name_list)):
        ws["A"+str(i+1)] = cn_en_name_list[i][0]
        ws["B"+str(i+1)] = cn_en_name_list[i][1]
    wb.save(excel_path)

    return cn_en_name_list

## 参数设置
root_dir = 'Stars'
man = False # True
# 每个人需要多少张图片
max_pic_num = 10
# 从第几个人开始 第一行就是序号为0
start = 24
# 指定下载图片尺寸类型
size_type = "extra large"
# 放大倍数
# b=1.5
# 裁剪后可接受的长宽比
# rate=1.3
#' 正脸 高清 大尺寸'
limit = ''
# 分别对应 全部尺寸0 特大尺寸9 大尺寸3 中尺寸2 小尺寸1 暂时没用到该参数 还是调上面self.url就行了
size_level = 9
# 分辨率分界线，低于这个的pass掉
resolution = 768
# 脸图比例，小于这个不考虑
face_img_rate = 0.03
# resize后的图像大小
img_size = 768

# 根据参数变化
gender = 'man' if man else 'woman'
excel_path = 'stars.xlsx'
sheet_num = 1 if man else 2
star_concept_name = 'concept'
star_output_name = 'generated' #若没有可自动创建

def main():
    create_dir_or_file(root_dir)
    # 中文名转为英文名，写入文件
    cn_en_write(excel_path)

    # 获取中英名list:[[cn,en],[cn,en],...]
    name_list=read_excel(excel_path)
    print('name_list',name_list)

    for I in tqdm(range(start,len(name_list))):
        cn_name,en_name=name_list[I][0],name_list[I][1]
        print('正在爬取第{}/{}:{}{}'.format(I,len(name_list),cn_name,en_name))

        # 为每个name创建文件夹
        en_name_dir=en_name
        
        os.makedirs(os.path.join(root_dir,gender,en_name_dir,star_concept_name), exist_ok=True)
        os.makedirs(os.path.join(root_dir,gender,en_name_dir,star_output_name), exist_ok=True)

        # 在concept中创建一个concept_word.txt，里面写上名字name_list[I][1]
        # path=os.path.join(root_dir,gender,en_name_dir,star_concept_name,'concept_word.txt')
        # with open(path,'w') as f:
        #     f.write(name_list[I][1])

        spider = Picture(I,cn_name,en_name,en_name_dir)
        spider.run()

if __name__ == '__main__':
    main()
